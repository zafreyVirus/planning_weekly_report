import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================
# CONFIGURATION
# ============================================

CSV_FOLDER = "data/"
OUTPUT_EXCEL = "planning_report_output.xlsx"

if os.path.exists(OUTPUT_EXCEL):
    try:
        with open(OUTPUT_EXCEL, 'a'):
            pass
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        OUTPUT_EXCEL = f"planning_report_output_{timestamp}.xlsx"
        print(f"⚠️ Original file was open. Creating new file: {OUTPUT_EXCEL}")

NE_NAMES = {
    "LMB": {
        "USN": "LMB_vUSN01",
        "CGW": "LMB_vCGW01",
        "DGW": "LMB_vDGW01",
        "UPCF": "dc2_LMBUPCF01",
        "MGW": "NEW LMBMGW",
        "VMGW": "LMBVMGW01",
        "VLR": "HMSCLMB_NEW"
    },
    "LL": {
        "USN": "CLOUDUSN",
        "CGW": "LLG_vCGW01",
        "DGW": "LLG_vDGW01",
        "UPCF": "dc1_LLGUPCF01",
        "MGW": "NEW LLGMGW",
        "VMGW": "LLWVMGW01",
        "VLR": "HMSCLLG_NEW"
    }
}

FILE_PATTERNS = {
    "2G_SAU_LMB": "USN SAU 2G LMB.csv",
    "2G_SAU_LL": "USN SAU 2G LLG.csv",
    "3G_SAU_LMB": "USN SAU 3G LMB.csv",
    "3G_SAU_LL": "USN SAU 3G LLG.csv",
    "4G_SAU_LMB": "USN SAU 4G LMB.csv",
    "4G_SAU_LL": "USN SAU 4G LLG.csv",
    "2G_PDP_LMB": "USN 2G PDP LMB.csv",
    "2G_PDP_LL": "USN 2G PDP LLG.csv",
    "3G_PDP_LMB": "vUSN 3G PDP LMB.csv",
    "3G_PDP_LL": "vUSN 3G PDP LLG.csv",
    "4G_PDP_LMB": "USN 4G PDP LMB.csv",
    "4G_PDP_LL": "USN 4G PDP LLG.csv",
    "CGW_2G3G_PDP": "CGW 2G3G PDP.csv",
    "CGW_4G_PDP": "CGW 4G PDP.csv",
    "DGW_THROUGHPUT_3G": "vDGW Throughput 3G.csv",
    "DGW_THROUGHPUT_4G": "vDGW Throughput 4G.csv",
    "UPCF_LMB": "UPCF PDP LMB.csv",
    "UPCF_LL": "UPCF PDP LLG.csv",
    "HLR": "HLR Subs.csv",
    "VLR": "VLR Subs.csv",
    "MGW": "MGW Traffic.csv",
    "VMGW": "vMGW Traffic.csv",
}

USN_METRIC_COLUMNS = {
    "2G_SAU": "Gb mode maximum attached users (number)",
    "3G_SAU": "Iu mode maximum attached users (number)",
    "4G_SAU": "Maximum attached users (number)",
    "2G_PDP": "Gb mode maximum act PDP context (number)",
    "3G_PDP": "Iu mode maximum active PDP context (number)",
    "4G_PDP": "Maximum bearer number in S1 mode (number)",
}

UGW_METRIC_COLUMNS = {
    "2G3G_PDP": "PGW-C 2/3G Maximum simultaneously activated PDP contexts (number)",
    "4G_PDP": "SGW-C and PGW-C combined Max number of Active EPS Bearers (number)",
    "THROUGHPUT_3G": "PGW-U 2/3G Gi downlink peak throughput in MB/s (MB/s)",
    "THROUGHPUT_4G": "User Plane SGi downlink user traffic peak throughput in MB/s (MB/s) (MB/s)",
}

UPCF_METRIC_COLUMN = "Maximum Number of Active Subscribers (number)"
HLR_METRIC_COLUMN = "Total Number of Subscribers (entries)"
VLR_METRIC_COLUMN = "Peak Number of VLR Subscribers (entries)"
MGW_METRIC_COLUMN = "Peak Licensed Traffic (number)"
VMGW_METRIC_COLUMN = "License Resource Usage for Basic Software on the MGW (%)"

LICENSE_CAPACITIES = {
    "HLR Subscribers (FE Users)": 7000000,
    "AuC Subscribers": 5300000,
    "HLR Subscribers (BE Users)": 10000000,
    "HSS LTE BE users": 10000000,
    "HSS LTE FE users": 7000000,
    "VLR LMB (new MSC)": 3300000,
    "VLR LLW (new MSC)": 3300000,
    "UMG TRAFFIC LMB": 30000,
    "UMG TRAFFIC LLW": 30000,
    "2G PDP(per sub) LIMBE": 2000,
    "2G PDP(per sub) LILONGWE": 2000,
    "2G SAU (per sub) LIMBE": 1500,
    "2G SAU (per K sub) LILONGWE": 1500,
    "3G PDP(per sub) LIMBE": 2000,
    "3G PDP(per sub) LILONGWE": 2000,
    "3G SAU (per sub) LIMBE": 1500,
    "3G SAU (per sub) LILONGWE": 1500,
    "4G PDP(per sub) LIMBE": 2000,
    "4G PDP(per sub) LILONGWE": 2000,
    "4G SAU (per sub) LIMBE": 1500,
    "4G SAU (per sub) LILONGWE": 1500,
    "UGW PDP 2G&3G (Per sub) LIMBE": 89000,
    "UGW PDP 4G (Per sub) LIMBE": 1018000,
    "UGW PDP 2G&3G (Per sub) LILONGWE": 388000,
    "UGW PDP 4G (Per sub) LILONGWE": 1436000,
    "UGW THROUGHPUT( MBps) LIMBE": 20000,
    "UGW THROUGHPUT( MBps) LILONGWE": 50000,
    "PCRF (PDP) LMB": 2000,
    "PCRF (PDP) LL": 2000,
}

HARDWARE_CAPACITIES = {
    "HLR BE Hardware": 10000000,
    "HLR FE Hardware": 7000000,
    "HSS LTE BE Hardware": 10000000,
    "HSS LTE FE Hardware": 7000000,
    "VLR Hardware": 3300000,
    "USN SAU Hardware": 1500,
    "USN PDP Hardware": 10000000,
    "UGW Throughput Hardware LMB": 20000,
    "UGW Throughput Hardware LL": 50000,
    "UPCF Hardware": 2000,
}

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
NE_CELL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALTERNATE_ROW_FILL = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


# ============================================
# HELPER FUNCTIONS
# ============================================

def find_data_start_row(file_path):
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
        for i, line in enumerate(lines):
            if 'Start Time' in line and 'Period (min)' in line:
                return i
    return 0


def read_csv_with_metadata(file_path):
    try:
        start_row = find_data_start_row(file_path)
        df = pd.read_csv(file_path, skiprows=start_row, encoding='utf-8-sig')
        df.columns = df.columns.str.strip('"')
        return df
    except Exception:
        return None


def get_max_from_last_7_days(df, ne_name, metric_column):
    if df is None or df.empty:
        return 0
    df_filtered = df[df['NE Name'] == ne_name]
    if df_filtered.empty:
        return 0
    df_filtered['Start Time'] = pd.to_datetime(df_filtered['Start Time'])
    max_date = df_filtered['Start Time'].max()
    cutoff_date = max_date - timedelta(days=7)
    df_last_7d = df_filtered[df_filtered['Start Time'] >= cutoff_date]
    if df_last_7d.empty:
        return 0
    return df_last_7d[metric_column].max()


def get_cloud_usn_value(region, metric_type):
    file_pattern_key = f"{metric_type}_{region}"
    if file_pattern_key not in FILE_PATTERNS:
        return 0
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS[file_pattern_key])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["USN"]
        metric_column = USN_METRIC_COLUMNS[metric_type]
        return get_max_from_last_7_days(df, ne_name, metric_column)
    except Exception:
        return 0


def get_cgw_value(region, metric_key):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS[f"CGW_{metric_key}"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["CGW"]
        metric_col = UGW_METRIC_COLUMNS[metric_key]
        return get_max_from_last_7_days(df, ne_name, metric_col)
    except Exception:
        return 0


def get_cgw_throughput_gi(region):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["DGW_THROUGHPUT_3G"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["DGW"]
        metric_col = UGW_METRIC_COLUMNS["THROUGHPUT_3G"]
        return get_max_from_last_7_days(df, ne_name, metric_col)
    except Exception:
        return 0


def get_cgw_throughput_sgi(region):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["DGW_THROUGHPUT_4G"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["DGW"]
        metric_col = UGW_METRIC_COLUMNS["THROUGHPUT_4G"]
        return get_max_from_last_7_days(df, ne_name, metric_col)
    except Exception:
        return 0


def get_upcf_value(region):
    file_pattern_key = f"UPCF_{region}"
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS[file_pattern_key])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        return get_max_from_last_7_days(df, NE_NAMES[region]["UPCF"], UPCF_METRIC_COLUMN)
    except Exception:
        return 0


def get_hlr_value():
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["HLR"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        return get_max_from_last_7_days(df, "HHLRLMB_USCDB", HLR_METRIC_COLUMN)
    except Exception:
        return 0


def get_vlr_value(region):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["VLR"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = "HMSCLMB_NEW" if region == "LMB" else "HMSCLLG_NEW"
        return get_max_from_last_7_days(df, ne_name, VLR_METRIC_COLUMN)
    except Exception:
        return 0


def get_mgw_value(region):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["MGW"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["MGW"]
        return get_max_from_last_7_days(df, ne_name, MGW_METRIC_COLUMN)
    except Exception:
        return 0


def get_vmgw_value(region):
    file_path = os.path.join(CSV_FOLDER, FILE_PATTERNS["VMGW"])
    if not os.path.exists(file_path):
        return 0
    try:
        df = read_csv_with_metadata(file_path)
        if df is None:
            return 0
        ne_name = NE_NAMES[region]["VMGW"]
        return get_max_from_last_7_days(df, ne_name, VMGW_METRIC_COLUMN)
    except Exception:
        return 0


# ============================================
# SHEET 1: CALCULATION
# ============================================

def generate_calculation_sheet():
    calculation_data = []

    # Cloud USN LMB
    for metric_display, metric_key in [
        ("vUSN 2G SAU (per k sub) LMB", "2G_SAU"),
        ("vUSN 3G SAU (per k sub) LMB", "3G_SAU"),
        ("vUSN 4G SAU (per k sub) LMB", "4G_SAU"),
        ("vUSN 2G PDP (per k sub) LMB", "2G_PDP"),
        ("vUSN 3G PDP (per k sub) LMB", "3G_PDP"),
        ("vUSN 4G PDP (per k sub) LMB", "4G_PDP"),
    ]:
        value = get_cloud_usn_value("LMB", metric_key)
        calculation_data.append({
            'Network Element': 'Cloud USN LMB',
            'Metric': metric_display,
            'vUSN LMB': value,
            'vUSN LLG': None,
            'Gi UGW (MB/s)': None,
            'SGi UGW (MB/s)': None,
        })

    # Cloud USN LLG
    for metric_display, metric_key in [
        ("vUSN 2G SAU (per k sub) LLG", "2G_SAU"),
        ("vUSN 3G SAU (per k sub) LLG", "3G_SAU"),
        ("vUSN 4G SAU (per k sub) LLG", "4G_SAU"),
        ("vUSN 2G PDP (per k sub) LLG", "2G_PDP"),
        ("vUSN 3G PDP (per k sub) LLG", "3G_PDP"),
        ("vUSN 4G PDP (per k sub) LLG", "4G_PDP"),
    ]:
        value = get_cloud_usn_value("LL", metric_key)
        calculation_data.append({
            'Network Element': 'Cloud USN LLG',
            'Metric': metric_display,
            'vUSN LMB': None,
            'vUSN LLG': value,
            'Gi UGW (MB/s)': None,
            'SGi UGW (MB/s)': None,
        })

    # Cloud UGW Limbe
    calculation_data.append({
        'Network Element': 'Cloud UGW Limbe',
        'Metric': 'vCGW 2G/3G PDP LMB',
        'vUSN LMB': None,
        'vUSN LLG': get_cgw_value("LMB", "2G3G_PDP"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'Cloud UGW Limbe',
        'Metric': 'vCGW 4G PDP LMB',
        'vUSN LMB': None,
        'vUSN LLG': get_cgw_value("LMB", "4G_PDP"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'Cloud UGW Limbe',
        'Metric': 'vDGW Throughput LMB',
        'vUSN LMB': None,
        'vUSN LLG': None,
        'Gi UGW (MB/s)': round(get_cgw_throughput_gi("LMB"), 2),
        'SGi UGW (MB/s)': round(get_cgw_throughput_sgi("LMB"), 2),
    })

    # Cloud UGW LLG
    calculation_data.append({
        'Network Element': 'Cloud UGW LLG',
        'Metric': 'vCGW 2G/3G PDP LLG',
        'vUSN LMB': None,
        'vUSN LLG': get_cgw_value("LL", "2G3G_PDP"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'Cloud UGW LLG',
        'Metric': 'vCGW 4G PDP LLG',
        'vUSN LMB': None,
        'vUSN LLG': get_cgw_value("LL", "4G_PDP"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'Cloud UGW LLG',
        'Metric': 'vDGW Throughput LLG',
        'vUSN LMB': None,
        'vUSN LLG': None,
        'Gi UGW (MB/s)': round(get_cgw_throughput_gi("LL"), 2),
        'SGi UGW (MB/s)': round(get_cgw_throughput_sgi("LL"), 2),
    })

    # UPCF
    calculation_data.append({
        'Network Element': 'UPCF',
        'Metric': 'UPCF PDP LMB',
        'vUSN LMB': get_upcf_value("LMB"),
        'vUSN LLG': None,
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'UPCF',
        'Metric': 'UPCF PDP LLG',
        'vUSN LMB': None,
        'vUSN LLG': get_upcf_value("LL"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })

    # MGW
    calculation_data.append({
        'Network Element': 'MGW',
        'Metric': 'Traffic (Erlangs) LMB',
        'vUSN LMB': get_mgw_value("LMB"),
        'vUSN LLG': None,
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'MGW',
        'Metric': 'Traffic (Erlangs) LLG',
        'vUSN LMB': None,
        'vUSN LLG': get_mgw_value("LL"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })

    # vMGW
    calculation_data.append({
        'Network Element': 'vMGW',
        'Metric': 'Traffic (Erlangs) LMB',
        'vUSN LMB': get_vmgw_value("LMB"),
        'vUSN LLG': None,
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })
    calculation_data.append({
        'Network Element': 'vMGW',
        'Metric': 'Traffic (Erlangs) LLG',
        'vUSN LMB': None,
        'vUSN LLG': get_vmgw_value("LL"),
        'Gi UGW (MB/s)': None,
        'SGi UGW (MB/s)': None,
    })

    return pd.DataFrame(calculation_data)


# ============================================
# SHEET 2: PROJECTIONS
# ============================================

def generate_projections_sheet():
    # Re-generate with full NE names for the lookup (before blanking)
    calculation_data = []

    for metric_display, metric_key in [
        ("vUSN 2G SAU (per k sub) LMB", "2G_SAU"),
        ("vUSN 3G SAU (per k sub) LMB", "3G_SAU"),
        ("vUSN 4G SAU (per k sub) LMB", "4G_SAU"),
        ("vUSN 2G PDP (per k sub) LMB", "2G_PDP"),
        ("vUSN 3G PDP (per k sub) LMB", "3G_PDP"),
        ("vUSN 4G PDP (per k sub) LMB", "4G_PDP"),
    ]:
        calculation_data.append({'Metric': metric_display, 'vUSN LMB': get_cloud_usn_value("LMB", metric_key), 'vUSN LLG': None})

    for metric_display, metric_key in [
        ("vUSN 2G SAU (per k sub) LLG", "2G_SAU"),
        ("vUSN 3G SAU (per k sub) LLG", "3G_SAU"),
        ("vUSN 4G SAU (per k sub) LLG", "4G_SAU"),
        ("vUSN 2G PDP (per k sub) LLG", "2G_PDP"),
        ("vUSN 3G PDP (per k sub) LLG", "3G_PDP"),
        ("vUSN 4G PDP (per k sub) LLG", "4G_PDP"),
    ]:
        calculation_data.append({'Metric': metric_display, 'vUSN LMB': None, 'vUSN LLG': get_cloud_usn_value("LL", metric_key)})

    calculation_data.append({'Metric': 'vCGW 2G/3G PDP LMB', 'vUSN LMB': None, 'vUSN LLG': get_cgw_value("LMB", "2G3G_PDP")})
    calculation_data.append({'Metric': 'vCGW 4G PDP LMB',    'vUSN LMB': None, 'vUSN LLG': get_cgw_value("LMB", "4G_PDP")})
    calculation_data.append({'Metric': 'vCGW 2G/3G PDP LLG', 'vUSN LMB': None, 'vUSN LLG': get_cgw_value("LL",  "2G3G_PDP")})
    calculation_data.append({'Metric': 'vCGW 4G PDP LLG',    'vUSN LMB': None, 'vUSN LLG': get_cgw_value("LL",  "4G_PDP")})
    calculation_data.append({'Metric': 'UPCF PDP LMB', 'vUSN LMB': get_upcf_value("LMB"), 'vUSN LLG': None})
    calculation_data.append({'Metric': 'UPCF PDP LLG', 'vUSN LMB': None, 'vUSN LLG': get_upcf_value("LL")})

    calc_df = pd.DataFrame(calculation_data)

    calc_lookup = {}
    for _, row in calc_df.iterrows():
        metric = row['Metric']
        if row['vUSN LMB'] is not None:
            calc_lookup[f"{metric}_LMB"] = row['vUSN LMB']
        if row['vUSN LLG'] is not None:
            calc_lookup[f"{metric}_LL"] = row['vUSN LLG']

    hlr_value = get_hlr_value()
    projections = [
        {'Metric': 'HLR Subscribers',        'Value': hlr_value},
        {'Metric': 'AuC Subscribers',         'Value': hlr_value},
        {'Metric': 'VLR LMB (new MSC)',       'Value': get_vlr_value('LMB')},
        {'Metric': 'VLR LLW (new MSC)',       'Value': get_vlr_value('LL')},
        {'Metric': 'UMG TRAFFIC LMB',         'Value': 0},
        {'Metric': 'UMG TRAFFIC LLW',         'Value': 0},
        {'Metric': '2G PDP(per sub) LIMBE',           'Value': calc_lookup.get('vUSN 2G PDP (per k sub) LMB_LMB', 0)},
        {'Metric': '2G PDP(per sub) LILONGWE',        'Value': calc_lookup.get('vUSN 2G PDP (per k sub) LLG_LL', 0)},
        {'Metric': '2G SAU (per sub) LIMBE',           'Value': calc_lookup.get('vUSN 2G SAU (per k sub) LMB_LMB', 0)},
        {'Metric': '2G SAU (per K sub) LILONGWE',      'Value': calc_lookup.get('vUSN 2G SAU (per k sub) LLG_LL', 0)},
        {'Metric': '3G PDP(per sub) LIMBE',            'Value': calc_lookup.get('vUSN 3G PDP (per k sub) LMB_LMB', 0)},
        {'Metric': '3G PDP(per sub) LILONGWE',         'Value': calc_lookup.get('vUSN 3G PDP (per k sub) LLG_LL', 0)},
        {'Metric': '3G SAU (per sub) LIMBE',            'Value': calc_lookup.get('vUSN 3G SAU (per k sub) LMB_LMB', 0)},
        {'Metric': '3G SAU (per sub) LILONGWE',         'Value': calc_lookup.get('vUSN 3G SAU (per k sub) LLG_LL', 0)},
        {'Metric': '4G PDP(per sub) LIMBE',             'Value': calc_lookup.get('vUSN 4G PDP (per k sub) LMB_LMB', 0)},
        {'Metric': '4G PDP(per sub) LILONGWE',          'Value': calc_lookup.get('vUSN 4G PDP (per k sub) LLG_LL', 0)},
        {'Metric': '4G SAU (per sub) LIMBE',             'Value': calc_lookup.get('vUSN 4G SAU (per k sub) LMB_LMB', 0)},
        {'Metric': '4G SAU (per sub) LILONGWE',          'Value': calc_lookup.get('vUSN 4G SAU (per k sub) LLG_LL', 0)},
        {'Metric': 'UGW PDP 2G&3G (Per sub) LIMBE',    'Value': calc_lookup.get('vCGW 2G/3G PDP LMB_LL', 0)},
        {'Metric': 'UGW PDP 4G (Per sub) LIMBE',       'Value': calc_lookup.get('vCGW 4G PDP LMB_LL', 0)},
        {'Metric': 'UGW PDP 2G&3G (Per sub) LILONGWE', 'Value': calc_lookup.get('vCGW 2G/3G PDP LLG_LL', 0)},
        {'Metric': 'UGW PDP 4G (Per sub) LILONGWE',    'Value': calc_lookup.get('vCGW 4G PDP LLG_LL', 0)},
        {'Metric': 'UGW THROUGHPUT( MBps) LIMBE',      'Value': get_cgw_throughput_gi("LMB") + get_cgw_throughput_sgi("LMB")},
        {'Metric': 'UGW THROUGHPUT( MBps) LILONGWE',   'Value': get_cgw_throughput_gi("LL")  + get_cgw_throughput_sgi("LL")},
        {'Metric': 'PCRF (PDP) LMB', 'Value': calc_lookup.get('UPCF PDP LMB_LMB', 0)},
        {'Metric': 'PCRF (PDP) LL',  'Value': calc_lookup.get('UPCF PDP LLG_LL', 0)},
    ]

    return pd.DataFrame(projections)


# ============================================
# SHEET 3: UTILIZATION PROJECTIONS
# ============================================

def generate_utilization_projections_sheet(projections_df):
    used_capacity = dict(zip(projections_df['Metric'], projections_df['Value']))

    metrics = [
        "HLR Subscribers (FE Users)", "AuC Subscribers", "HLR Subscribers (BE Users)",
        "HSS LTE BE users", "HSS LTE FE users", "VLR LMB (new MSC)", "VLR LLW (new MSC)",
        "UMG TRAFFIC LMB", "UMG TRAFFIC LLW", "2G PDP(per sub) LIMBE", "2G PDP(per sub) LILONGWE",
        "2G SAU (per sub) LIMBE", "2G SAU (per K sub) LILONGWE", "3G PDP(per sub) LIMBE",
        "3G PDP(per sub) LILONGWE", "3G SAU (per sub) LIMBE", "3G SAU (per sub) LILONGWE",
        "4G PDP(per sub) LIMBE", "4G PDP(per sub) LILONGWE", "4G SAU (per sub) LIMBE",
        "4G SAU (per sub) LILONGWE", "UGW PDP 2G&3G (Per sub) LIMBE", "UGW PDP 4G (Per sub) LIMBE",
        "UGW PDP 2G&3G (Per sub) LILONGWE", "UGW PDP 4G (Per sub) LILONGWE",
        "UGW THROUGHPUT( MBps) LIMBE", "UGW THROUGHPUT( MBps) LILONGWE",
        "PCRF (PDP) LMB", "PCRF (PDP) LL",
    ]

    utilization_data = []
    for metric in metrics:
        used = used_capacity.get(metric, 0)
        license_capacity = LICENSE_CAPACITIES.get(metric, 1)
        utilization_pct = (used / license_capacity) * 100 if license_capacity > 0 else 0
        utilization_data.append({
            'Metric': metric,
            'Used Capacity': used,
            'License Capacity': license_capacity,
            'Utilization (%)': round(utilization_pct, 2)
        })

    return pd.DataFrame(utilization_data)


# ============================================
# SHEET 4: HARDWARE UTILIZATION
# ============================================

def generate_hardware_utilization_sheet(projections_df):
    used_capacity = dict(zip(projections_df['Metric'], projections_df['Value']))

    hardware_items = [
        ('HLR BE Hardware',             'HLR Subscribers (BE Users)'),
        ('HLR FE Hardware',             'HLR Subscribers (FE Users)'),
        ('HSS LTE BE Hardware',         'HSS LTE BE users'),
        ('HSS LTE FE Hardware',         'HSS LTE FE users'),
        ('VLR Hardware',                'VLR LMB (new MSC)'),
        ('USN SAU Hardware',            '2G SAU (per sub) LIMBE'),
        ('USN PDP Hardware',            '2G PDP(per sub) LIMBE'),
        ('UGW Throughput Hardware LMB', 'UGW THROUGHPUT( MBps) LIMBE'),
        ('UGW Throughput Hardware LL',  'UGW THROUGHPUT( MBps) LILONGWE'),
        ('UPCF Hardware',               'PCRF (PDP) LMB'),
    ]

    hardware_data = []
    for hw_metric, proj_metric in hardware_items:
        used = used_capacity.get(proj_metric, 0)
        capacity = HARDWARE_CAPACITIES.get(hw_metric, 1)
        utilization = (used / capacity) * 100 if capacity > 0 else 0
        hardware_data.append({
            'Hardware Component': hw_metric,
            'Used Capacity': used,
            'Hardware Capacity': capacity,
            'Utilization (%)': round(utilization, 2)
        })

    return pd.DataFrame(hardware_data)


# ============================================
# SHEET 5: HARDWARE REDUNDANCY
# ============================================

def generate_hardware_redundancy_sheet():
    return pd.DataFrame([
        {'Component': 'HLR',  'Active': 1, 'Standby': 1, 'Total': 2, 'Redundancy Type': '1+1'},
        {'Component': 'HSS',  'Active': 1, 'Standby': 1, 'Total': 2, 'Redundancy Type': '1+1'},
        {'Component': 'VLR',  'Active': 2, 'Standby': 1, 'Total': 3, 'Redundancy Type': 'N+1'},
        {'Component': 'USN',  'Active': 2, 'Standby': 1, 'Total': 3, 'Redundancy Type': 'N+1'},
        {'Component': 'UGW',  'Active': 2, 'Standby': 1, 'Total': 3, 'Redundancy Type': 'N+1'},
        {'Component': 'UPCF', 'Active': 1, 'Standby': 1, 'Total': 2, 'Redundancy Type': '1+1'},
    ])


# ============================================
# SHEET 6: CURRENT COMMERCIAL LICENSE
# ============================================

def generate_commercial_license_sheet():
    return pd.DataFrame([
        {'Component': 'HLR FE License',             'License Quantity': 7000000,  'Unit': 'Users'},
        {'Component': 'HLR BE License',             'License Quantity': 10000000, 'Unit': 'Users'},
        {'Component': 'HSS LTE FE License',         'License Quantity': 7000000,  'Unit': 'Users'},
        {'Component': 'HSS LTE BE License',         'License Quantity': 10000000, 'Unit': 'Users'},
        {'Component': 'VLR License',                'License Quantity': 3300000,  'Unit': 'Users'},
        {'Component': 'USN SAU License',            'License Quantity': 1500,     'Unit': 'Per K sub'},
        {'Component': 'USN PDP License',            'License Quantity': 10000000, 'Unit': 'Users'},
        {'Component': 'UGW Throughput License LMB', 'License Quantity': 20000,    'Unit': 'MB/s'},
        {'Component': 'UGW Throughput License LL',  'License Quantity': 50000,    'Unit': 'MB/s'},
        {'Component': 'UPCF License',               'License Quantity': 2000,     'Unit': 'Per K sub'},
    ])


# ============================================
# SHEET 7: COMMERCIAL LICENSE REDUNDANCY
# ============================================

def generate_commercial_license_redundancy_sheet(commercial_df):
    redundancy_data = []
    for _, row in commercial_df.iterrows():
        qty = row['License Quantity']
        redundancy_data.append({
            'Component':       row['Component'],
            'Active License':  qty,
            'Standby License': qty,
            'Total License':   qty * 2,
            'Redundancy Factor': '2x'
        })
    return pd.DataFrame(redundancy_data)


# ============================================
# STYLING FUNCTIONS
# ============================================

def write_calculation_sheet(workbook, df_calc):
    """Write the Calculation sheet directly via openpyxl — no to_excel().
    This gives us full control over row layout so NE merging works reliably.
    """
    ws = workbook.create_sheet('Calculation', 0)

    column_widths = {'A': 25, 'B': 38, 'C': 18, 'D': 18, 'E': 18, 'F': 18}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ── Row 1: Title ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    title_cell = ws.cell(row=1, column=1, value="Network Element Capacity Calculations")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = SECTION_HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # ── Row 2: Column headers ──────────────────────────────────────────────
    headers = ['Network Element', 'Metric', 'vUSN LMB', 'vUSN LLG', 'Gi UGW (MB/s)', 'SGi UGW (MB/s)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    # ── Rows 3+: Data — identify groups while writing ──────────────────────
    # We track NE groups here directly from the DataFrame (NE values are intact).
    cols = ['Network Element', 'Metric', 'vUSN LMB', 'vUSN LLG', 'Gi UGW (MB/s)', 'SGi UGW (MB/s)']
    data_start_row = 3

    # First pass: build group list
    groups = []
    current_ne = None
    group_start = data_start_row
    for i, row in df_calc.iterrows():
        excel_row = data_start_row + i
        ne = row['Network Element']
        if ne != current_ne:
            if current_ne is not None:
                groups.append({'name': current_ne, 'start': group_start, 'end': excel_row - 1})
            current_ne = ne
            group_start = excel_row
    if current_ne is not None:
        groups.append({'name': current_ne, 'start': group_start, 'end': data_start_row + len(df_calc) - 1})

    # Second pass: write all data cells
    for i, row in df_calc.iterrows():
        excel_row = data_start_row + i
        is_alt = i % 2 == 0  # alternating based on data row index

        for col_idx, col_name in enumerate(cols, start=1):
            val = row[col_name] if col_name in row else None
            # Skip NE column for now — handled via merge below
            if col_idx == 1:
                continue
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = None if (val is None or (isinstance(val, float) and pd.isna(val))) else val
            cell.border = BORDER
            cell.alignment = LEFT_ALIGN if col_idx == 2 else RIGHT_ALIGN
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if col_idx >= 5 else '#,##0'
            if is_alt:
                cell.fill = ALTERNATE_ROW_FILL

    # Third pass: write + merge Network Element column
    for grp in groups:
        s, e = grp['start'], grp['end']
        # Style all rows in the group for col A border/fill first
        for r in range(s, e + 1):
            c = ws.cell(row=r, column=1)
            c.fill = NE_CELL_FILL
            c.border = BORDER
        # Merge if multi-row
        if s < e:
            ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
        # Set value and style on the top (anchor) cell
        anchor = ws.cell(row=s, column=1)
        anchor.value = grp['name']
        anchor.font = Font(bold=True, size=11, color="1a4a6e")
        anchor.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        anchor.fill = NE_CELL_FILL
        anchor.border = BORDER


def apply_general_styling(worksheet, title, has_value_column=True):
    """Apply general styling to any sheet."""

    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 35)

    header_row = 1
    for cell in worksheet[header_row]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        for col_idx, cell in enumerate(worksheet[row_idx], start=1):
            cell.border = BORDER
            if col_idx == 1:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = RIGHT_ALIGN
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            if row_idx % 2 == 0:
                cell.fill = ALTERNATE_ROW_FILL

    worksheet.insert_rows(1)
    title_cell = worksheet['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = SECTION_HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    num_cols = max(1, worksheet.max_column)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    worksheet.row_dimensions[1].height = 28


def style_utilization_sheet(worksheet):
    """Utilization Projections with red/green color coding."""

    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20

    for col_idx, header in enumerate(['Metric', 'Used Capacity', 'License Capacity', 'Utilization (%)'], start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.alignment = LEFT_ALIGN if col_idx == 1 else RIGHT_ALIGN
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0' if col_idx < 4 else '0.00'
            if col_idx == 4:
                try:
                    val = float(cell.value)
                    if val > 80:
                        cell.fill = WARNING_FILL
                        cell.font = Font(bold=True, color="9C0006")
                    elif val > 60:
                        cell.fill = GOOD_FILL
                    elif row_idx % 2 == 0:
                        cell.fill = ALTERNATE_ROW_FILL
                except Exception:
                    pass
            elif row_idx % 2 == 0:
                cell.fill = ALTERNATE_ROW_FILL

    worksheet.insert_rows(1)
    title_cell = worksheet['A1']
    title_cell.value = "Utilization Projections"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = SECTION_HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    worksheet.row_dimensions[1].height = 28


# ============================================
# MAIN
# ============================================

if __name__ == "__main__":

    if not os.path.exists(CSV_FOLDER):
        print(f"Creating folder: {CSV_FOLDER}")
        os.makedirs(CSV_FOLDER)
        print(f"Please place your CSV files in '{CSV_FOLDER}' and run again.")
    else:
        print("=" * 60)
        print("GENERATING COMPLETE EXCEL REPORT WITH 7 SHEETS")
        print("=" * 60)

        print("\n1. Generating Calculation sheet...")
        df_calc = generate_calculation_sheet()

        print("2. Generating Projections sheet...")
        df_projections = generate_projections_sheet()

        print("3. Generating Utilization Projections sheet...")
        df_utilization = generate_utilization_projections_sheet(df_projections)

        print("4. Generating Hardware Utilization sheet...")
        df_hardware_util = generate_hardware_utilization_sheet(df_projections)

        print("5. Generating Hardware Redundancy sheet...")
        df_hardware_red = generate_hardware_redundancy_sheet()

        print("6. Generating Current Commercial License sheet...")
        df_commercial = generate_commercial_license_sheet()

        print("7. Generating Commercial License Redundancy sheet...")
        df_commercial_red = generate_commercial_license_redundancy_sheet(df_commercial)

        # Write all sheets
        print("\n8. Writing and styling sheets...")
        from openpyxl import Workbook as OpenpyxlWorkbook

        # Step A: Write Calculation sheet directly via openpyxl and save
        wb = OpenpyxlWorkbook()
        wb.remove(wb.active)
        write_calculation_sheet(wb, df_calc)
        wb.save(OUTPUT_EXCEL)

        # Step B: Append remaining sheets using pandas (mode='a', keep_vba not needed)
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl', mode='a') as writer:
            df_projections.to_excel(writer,    sheet_name='Projections',                   index=False)
            df_utilization.to_excel(writer,    sheet_name='Utilization Projections',       index=False)
            df_hardware_util.to_excel(writer,  sheet_name='Hardware Utilization',          index=False)
            df_hardware_red.to_excel(writer,   sheet_name='Hardware Redundancy',           index=False)
            df_commercial.to_excel(writer,     sheet_name='Current Commercial License',    index=False)
            df_commercial_red.to_excel(writer, sheet_name='Commercial License Redundancy', index=False)

        # Step C: Style the appended sheets
        wb = load_workbook(OUTPUT_EXCEL)
        apply_general_styling(wb['Projections'],                   "Projections Input")
        style_utilization_sheet(wb['Utilization Projections'])
        apply_general_styling(wb['Hardware Utilization'],          "Hardware Utilization")
        apply_general_styling(wb['Hardware Redundancy'],           "Hardware Redundancy Configuration",        has_value_column=False)
        apply_general_styling(wb['Current Commercial License'],    "Current Commercial License Inventory",     has_value_column=False)
        apply_general_styling(wb['Commercial License Redundancy'], "Commercial License Redundancy Configuration", has_value_column=False)

        wb.save(OUTPUT_EXCEL)

        print("\n" + "=" * 60)
        print(f"✅ Complete Excel report saved to: {OUTPUT_EXCEL}")
        print("=" * 60)

        print("\n📊 CALCULATION SHEET PREVIEW:")
        print(df_calc.to_string(index=False))