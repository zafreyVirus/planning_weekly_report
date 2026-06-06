"""
planning_report_writer.py
Writes the Calculation sheet in the same layout/style as the manual planning report.
Call write_calculation_sheet(data, output_path) with the dict from generate_calculation_sheet().
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PINK_FILL = PatternFill("solid", fgColor="CCECFF")
BLUE_FILL = PatternFill("solid", fgColor="C2EDFA")
NF        = '#,##0'
NF_DEC    = '#,##0.00'


def _border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)


def _c(ws, coord, value, fill=None, bold=False, align="left", num_format=None):
    cell = ws[coord]
    cell.value = value
    if fill:
        cell.fill = fill
    cell.font = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _border()
    if num_format:
        cell.number_format = num_format


def write_calculation_sheet(data: dict, output_path: str):
    """
    data keys expected:
        usn_lmb, usn_ll  -> dict with keys: 2G SAU, 3G SAU, 4G SAU, 2G PDP, 3G PDP, 4G PDP
        ugw_lmb, ugw_ll  -> dict with keys: 2G3G PDP, 4G PDP, Throughput
        upcf_lmb, upcf_ll -> scalar
        mgw_lmb, mgw_ll   -> scalar
        vmgw_lmb, vmgw_ll -> scalar (percent)
    """
    usn_lmb  = data["usn_lmb"]
    usn_ll   = data["usn_ll"]
    ugw_lmb  = data["ugw_lmb"]
    ugw_ll   = data["ugw_ll"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation"

    col_widths = {"A": 18, "B": 32, "C": 16, "D": 16, "E": 16, "F": 12, "G": 12, "H": 16, "I": 16}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 2-3 headers
    for col, val in {"F": "Gi", "G": "Sgi"}.items():
        _c(ws, f"{col}2", val, align="center")
    for col, val in {"D": "USN", "E": "vUSN LLG", "F": "UGW", "G": "UGW", "H": "vUSN LMB", "I": "Total"}.items():
        _c(ws, f"{col}3", val, align="center")

    def row(r, ne, kpi, d_val, c_expr, e_val=None, i_expr=None, kpi_fill=None, numf=NF):
        _c(ws, f"A{r}", ne)
        _c(ws, f"B{r}", kpi, fill=kpi_fill or PINK_FILL)
        _c(ws, f"C{r}", c_expr, fill=PINK_FILL, num_format=numf)
        _c(ws, f"D{r}", d_val, num_format=numf)
        if e_val is not None:
            _c(ws, f"E{r}", e_val, num_format=numf)
        if i_expr:
            _c(ws, f"I{r}", i_expr, fill=PINK_FILL, num_format=numf)

    # Cloud USN LMB (rows 4-9)
    usn_lmb_kpis = [
        ("vUSN 2G SAU (per k sub) LMB", "2G SAU", BLUE_FILL),
        ("vUSN 3G SAU (per k sub) LMB", "3G SAU", BLUE_FILL),
        ("vUSN 4G SAU (per k sub) LMB", "4G SAU", BLUE_FILL),
        ("vUSN 2G PDP (per k sub) LMB", "2G PDP", None),
        ("vUSN 3G PDP (per k sub) LMB", "3G PDP", None),
        ("vUSN 4G PDP (per k sub) LMB", "4G PDP", None),
    ]
    for i, (kpi, key, fill) in enumerate(usn_lmb_kpis, start=4):
        ne = "Cloud USN" if i == 4 else ""
        row(i, ne, kpi, usn_lmb[key], f"=D{i}", i_expr=f"=C{i}", kpi_fill=fill)

    # Cloud USN LL (rows 10-15)
    usn_ll_kpis = [
        ("vUSN 2G SAU (per k sub) LL", "2G SAU", BLUE_FILL),
        ("vUSN 3G SAU (per k sub) LL", "3G SAU", BLUE_FILL),
        ("vUSN 4G SAU (per k sub) LL", "4G SAU", BLUE_FILL),
        ("vUSN 2G PDP (per k sub) LL", "2G PDP", None),
        ("vUSN 3G PDP (per k sub) LL", "3G PDP", None),
        ("vUSN 4G PDP (per k sub) LL", "4G PDP", None),
    ]
    for i, (kpi, key, fill) in enumerate(usn_ll_kpis, start=10):
        ne = "Cloud USN" if i == 10 else ""
        row(i, ne, kpi, usn_ll[key], f"=E{i}", e_val=usn_ll[key], i_expr=f"=C{i}", kpi_fill=fill)

    # Cloud UGW LMB (rows 16-18)
    ugw_lmb_data = [
        ("vDGW 2G/3G PDP LMB",  "2G3G PDP", NF),
        ("vDGW 4G PDP LMB",     "4G PDP",   NF),
        ("vDGW Throughput LMB", "Throughput", NF_DEC),
    ]
    for i, (kpi, key, numf) in enumerate(ugw_lmb_data, start=16):
        ne = "Cloud UGW Limbe" if i == 16 else ""
        row(i, ne, kpi, ugw_lmb[key], f"=D{i}", i_expr=f"=C{i}", numf=numf)

    # Cloud UGW LLG (rows 19-21)
    ugw_ll_data = [
        ("vDGW 2G/3G PDP LLW",  "2G3G PDP", NF),
        ("vDGW 4G PDP LLW",     "4G PDP",   NF),
        ("vDGW Throughput LLW", "Throughput", NF_DEC),
    ]
    for i, (kpi, key, numf) in enumerate(ugw_ll_data, start=19):
        ne = "Cloud UGW LLG" if i == 19 else ""
        row(i, ne, kpi, ugw_ll[key], f"=D{i}", i_expr=f"=C{i}", numf=numf)

    # MGW (rows 22-23)
    for i, (ne, kpi, val) in enumerate([
        ("MGW", "Traffic (Erlangs) LMB", data["mgw_lmb"]),
        ("",    "Traffic (Erlangs) LL",  data["mgw_ll"]),
    ], start=22):
        row(i, ne, kpi, val, f"=D{i}", i_expr=f"=C{i}")

    # vMGW (rows 24-25)
    for i, (ne, kpi, val) in enumerate([
        ("vMGW", "License Usage (%) LMB", data["vmgw_lmb"]),
        ("",     "License Usage (%) LL",  data["vmgw_ll"]),
    ], start=24):
        row(i, ne, kpi, val, f"=D{i}", i_expr=f"=C{i}")

    # UPCF (rows 26-27)
    for i, (ne, kpi, val) in enumerate([
        ("UPCF", "UPCF PDP LMB", data["upcf_lmb"]),
        ("",     "UPCF PDP LLG", data["upcf_ll"]),
    ], start=26):
        row(i, ne, kpi, val, f"=D{i}", i_expr=f"=C{i}")

    for r in range(4, 28):
        ws.row_dimensions[r].height = 18

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")


# ---- Integration shim: build data dict from generate_calculation_sheet() results ----
def build_data_from_df(df):
    """Convert the DataFrame from generate_calculation_sheet() into the dict format."""
    def v(ne_type, region, metric):
        row = df[(df['NE Type'] == ne_type) & (df['Region'] == region) & (df['Metric'] == metric)]
        return float(row['Value'].values[0]) if len(row) else 0

    return {
        "usn_lmb": {
            "2G SAU": v("Cloud USN", "LMB", "2G SAU (per K sub)"),
            "3G SAU": v("Cloud USN", "LMB", "3G SAU (per K sub)"),
            "4G SAU": v("Cloud USN", "LMB", "4G SAU (per K sub)"),
            "2G PDP": v("Cloud USN", "LMB", "2G PDP (per K sub)"),
            "3G PDP": v("Cloud USN", "LMB", "3G PDP (per K sub)"),
            "4G PDP": v("Cloud USN", "LMB", "4G PDP (per K sub)"),
        },
        "usn_ll": {
            "2G SAU": v("Cloud USN", "LL", "2G SAU (per K sub)"),
            "3G SAU": v("Cloud USN", "LL", "3G SAU (per K sub)"),
            "4G SAU": v("Cloud USN", "LL", "4G SAU (per K sub)"),
            "2G PDP": v("Cloud USN", "LL", "2G PDP (per K sub)"),
            "3G PDP": v("Cloud USN", "LL", "3G PDP (per K sub)"),
            "4G PDP": v("Cloud USN", "LL", "4G PDP (per K sub)"),
        },
        "ugw_lmb": {
            "2G3G PDP":  v("Cloud UGW", "LMB", "vCGW 2G/3G PDP"),
            "4G PDP":    v("Cloud UGW", "LMB", "vCGW 4G PDP"),
            "Throughput": v("Cloud UGW", "LMB", "vCGW Throughput (MB/s)"),
        },
        "ugw_ll": {
            "2G3G PDP":  v("Cloud UGW", "LL", "vCGW 2G/3G PDP"),
            "4G PDP":    v("Cloud UGW", "LL", "vCGW 4G PDP"),
            "Throughput": v("Cloud UGW", "LL", "vCGW Throughput (MB/s)"),
        },
        "upcf_lmb": v("UPCF", "LMB", "Maximum Active Subscribers"),
        "upcf_ll":  v("UPCF", "LL",  "Maximum Active Subscribers"),
        "mgw_lmb":  v("MGW",  "LMB", "Peak Licensed Traffic"),
        "mgw_ll":   v("MGW",  "LL",  "Peak Licensed Traffic"),
        "vmgw_lmb": v("vMGW", "LMB", "License Usage (%)"),
        "vmgw_ll":  v("vMGW", "LL",  "License Usage (%)"),
    }